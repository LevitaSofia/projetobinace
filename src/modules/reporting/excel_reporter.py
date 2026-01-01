
import sqlite3
import pandas as pd
import json
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_FILE = '/home/ubuntu/projetobinace/sandra_trading.db'
OUTPUT_FILE = '/home/ubuntu/projetobinace/relatorio_sandra.xlsx'
USDT_BRL_RATE = 6.10 

# --- GOOGLE SHEETS CONFIG ---
GSHEET_ID = "19wkbjAT4L6KPXXSRexEEbMs95nsAjNwglDZx81bllnw"
CREDENTIALS_FILE = '/home/ubuntu/projetobinace/credentials.json'

try:
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    HAS_GSHEETS = True
except ImportError:
    HAS_GSHEETS = False
    print("⚠️ Bibliotecas gspread/oauth2client não encontradas. GSheets desativado.")

# --- BINANCE DATA ---
try:
    import ccxt
    from dotenv import load_dotenv
    load_dotenv()
    API_KEY = os.getenv('BINANCE_API_KEY')
    SECRET = os.getenv('BINANCE_SECRET')
    HAS_BINANCE = True if API_KEY and SECRET else False
except ImportError:
    HAS_BINANCE = False

def get_live_usd_brl(exchange):
    """Tenta pegar cotação USDT/BRL da Binance."""
    try:
        ticker = exchange.fetch_ticker('USDT/BRL')
        price = float(ticker['last'])
        print(f"💵 Cotação USDT/BRL Atual: R$ {price:.2f}")
        return price
    except Exception as e:
        print(f"⚠️ Erro cotação BRL: {e}")
        return 6.10

def get_last_trade_data(symbol):
    """Busca dados completos da última compra desse par no banco."""
    if not os.path.exists(DB_FILE):
        return {}
    
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT json_data FROM trade_history WHERE symbol=? AND side='buy' ORDER BY timestamp DESC LIMIT 1", 
            (symbol,)
        )
        row = cursor.fetchone()
        conn.close()
        
        if row:
            return json.loads(row[0])
    except:
        pass
    return {}

def get_binance_portfolio():
    """Busca saldo e posições abertas na Binance."""
    if not HAS_BINANCE:
        return 0, []
    
    try:
        exchange = ccxt.binance({
            'apiKey': API_KEY,
            'secret': SECRET,
            'enableRateLimit': True
        })
        
        # Pega taxa BRL real time
        brl_rate = get_live_usd_brl(exchange)
        global USDT_BRL_RATE
        USDT_BRL_RATE = brl_rate

        balance = exchange.fetch_balance()
        total_balance_usdt = float(balance.get('total', {}).get('USDT', 0))
        
        tickers = exchange.fetch_tickers()
        
        holdings = []
        estimated_total_wallet = total_balance_usdt 
        
        # Itera sobre moedas com saldo
        for coin, qtd in balance['total'].items():
            if qtd > 0 and coin != 'USDT':
                symbol = f"{coin}/USDT"
                current_price = 0.0
                if symbol in tickers:
                    current_price = float(tickers[symbol]['last'])
                
                # Ignora poeira e moedas sem valor
                value_usdt = qtd * current_price
                if value_usdt < 0.5: 
                    continue

                estimated_total_wallet += value_usdt
                
                # Busca dados históricos (preço, tempo, previsões)
                trade_data = get_last_trade_data(symbol)
                entry_price = float(trade_data.get('price', 0.0))
                
                # Se não achou no banco, usa placeholder
                if entry_price == 0:
                    entry_price = current_price

                pnl_usdt = (current_price - entry_price) * qtd
                pnl_pct = ((current_price - entry_price) / entry_price) * 100 if entry_price > 0 else 0.0

                # Previsões (Fallback para trades antigos)
                pred_profit = float(trade_data.get('pred_profit', 0.0))
                est_mins = int(trade_data.get('est_duration', 0))
                buy_time_str = trade_data.get('timestamp')
                
                # Se não tiver previsão gravada (trade antigo), estima agora
                if pred_profit == 0 and entry_price > 0:
                     pred_profit = (entry_price * qtd) * 0.015 # 1.5% alvo padrão
                
                if est_mins == 0:
                     est_mins = 240 # 4 horas padrão
                
                time_info = "N/A"
                if buy_time_str:
                    try:
                        buy_dt = datetime.fromisoformat(buy_time_str)
                        target_dt = buy_dt + timedelta(minutes=est_mins)
                        now = datetime.now()
                        remaining = target_dt - now
                        
                        if remaining.total_seconds() > 0:
                            hours_left = int(remaining.total_seconds() // 3600)
                            time_info = f"{hours_left}h restantes"
                        else:
                            time_info = "Vencido (Atrasado)"
                    except:
                        pass
                else:
                    # Se não tem timestamp (compra manual/externa?), assume 4h a partir de agora
                    time_info = "~4h (Est.)"

                holdings.append({
                    'Moeda': coin,
                    'Par': symbol,
                    'Quantidade': qtd,
                    'Preço Compra ($)': entry_price,
                    'Preço Atual ($)': current_price,
                    'Total Investido ($)': entry_price * qtd,
                    'Valor Atual ($)': value_usdt,
                    'PnL Estimado ($)': pnl_usdt,
                    'PnL Estimado (%)': pnl_pct / 100.0,
                    'Lucro Alvo ($)': pred_profit,
                    'Tempo Est.': time_info,
                    'Valor em Reais (R$)': value_usdt * brl_rate
                })
        
        # Adiciona USDT
        if total_balance_usdt > 0:
            holdings.append({
                'Moeda': 'USDT',
                'Par': 'Dólar',
                'Quantidade': total_balance_usdt,
                'Preço Compra ($)': 1.0,
                'Preço Atual ($)': 1.0,
                'Total Investido ($)': total_balance_usdt,
                'Valor Atual ($)': total_balance_usdt,
                'PnL Estimado ($)': 0.0,
                'PnL Estimado (%)': 0.0,
                'Lucro Alvo ($)': 0.0,
                'Tempo Est.': '-',
                'Valor em Reais (R$)': total_balance_usdt * brl_rate
            })

        return estimated_total_wallet, holdings
    except Exception as e:
        print(f"⚠️ Erro ao ler Binance: {e}")
        return 0, []

def generate_report():
    if not os.path.exists(DB_FILE):
        print(f"Banco de dados não encontrado: {DB_FILE}")
        return

    try:
        # 1. Busca Dados da Binance (Carteira Atual)
        total_wallet_val, portfolio_list = get_binance_portfolio()
        
        conn = sqlite3.connect(DB_FILE)
        # ORDENAÇÃO CORRIGIDA: Mais recentes primeiro (DESC)
        query = "SELECT json_data FROM trade_history WHERE side='sell' ORDER BY timestamp DESC"
        cursor = conn.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        conn.close()

        data_extrato = []
        cumulative_profit = 0.0
        
        # Precisamos inverter a lista para calcular o cumulativo corretamente (do passado pro futuro)
        # mas exibir do futuro pro passado.
        # Estratégia: Calcular cumulativo na ordem ASC, depois inverter a lista para display.
        
        temp_rows = []
        if rows:
            for row in rows:
                temp_rows.append(json.loads(row[0]))
        
        # Ordena cronologicamente para calcular acumulado
        temp_rows.sort(key=lambda x: x.get('timestamp', ''))
        
        processed_data = []
        total_profit = 0.0
        
        for t in temp_rows:
            ts = t.get('timestamp', '')
            try:
                dt_obj = datetime.fromisoformat(ts)
                date_str = dt_obj.strftime('%d/%m/%Y')
                time_str = dt_obj.strftime('%H:%M:%S')
                month_key = dt_obj.strftime('%Y-%m')
            except:
                date_str, time_str, month_key = ts, '', 'Outros'

            symbol = t.get('symbol')
            entry_price = float(t.get('entry_price', 0))
            exit_price = float(t.get('exit_price', 0))
            qty = float(t.get('qty', 0))
            fees = float(t.get('fees', 0))
            profit_usdt = float(t.get('net_profit_usdt', 0))
            profit_pct = float(t.get('net_profit_pct', 0))
            # Extração de Indicadores Técnicos
            rsi_val = float(t.get('rsi', 0))
            adx_val = float(t.get('adx', 0))
            vol_ratio = float(t.get('vol_ratio', 0))
            
            # Fibonacci info
            fib_info = t.get('fibonacci', {})
            fib_level = fib_info.get('touching_level', '') if isinstance(fib_info, dict) else ''
            
            ml_prob = t.get('ml_prob') 
            
            ml_prob_display = 0.0
            if ml_prob is not None:
                ml_prob_display = float(ml_prob) / 100.0 # Decimal para porcentagem Excel
            else:
                ml_prob_display = 0.0 # Sem dados

            invested = entry_price * qty
            sold_total = exit_price * qty
            
            total_profit += profit_usdt
            
            processed_data.append({
                'Mês': month_key,
                'Data': date_str,
                'Hora': time_str,
                'Moeda': symbol,
                'Tipo': 'VENDA (Lucro Realizado)',
                'Preço Compra': entry_price,
                'Preço Venda': exit_price,
                'Qtd': qty,
                'Investido ($)': invested,
                'Retorno ($)': sold_total,
                'Lucro ($)': profit_usdt,
                'Lucro (%)': profit_pct / 100.0,
                'Acumulado ($)': total_profit,
                'RSI': rsi_val,
                'ADX': adx_val,
                'Vol Ratio': vol_ratio,
                'Fibo Level': fib_level,
                '🤖 Prob. ML': ml_prob_display
            })

        # Agora inverte para o mais recente ficar em cima
        processed_data.reverse()

        df_extrato = pd.DataFrame(processed_data)
        df_portfolio = pd.DataFrame(portfolio_list)

        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
            # Aba 1: Carteira Atual (Privilegiada)
            if not df_portfolio.empty:
                df_portfolio.to_excel(writer, sheet_name='Carteira Atual', index=False)
            
            # Aba 2: Extrato Geral
            if not df_extrato.empty:
                df_extrato.drop(columns=['Mês']).to_excel(writer, sheet_name='Extrato Geral', index=False)
            
            # Aba 3: Performance por Moeda
            if not df_extrato.empty:
                pivot = df_extrato.groupby('Moeda')[['Investido ($)', 'Lucro ($)']].sum()
                pivot['ROI (%)'] = pivot['Lucro ($)'] / pivot['Investido ($)'] if pivot['Investido ($)'].sum() != 0 else 0
                pivot = pivot.sort_values('Lucro ($)', ascending=False)
                pivot['ROI (%)'] = pivot['Lucro ($)'] / pivot['Investido ($)'] if pivot['Investido ($)'].sum() != 0 else 0
                pivot = pivot.sort_values('Lucro ($)', ascending=False)
                pivot.to_excel(writer, sheet_name='Performance')

            # Aba 4: 📘 Como Funciona ML (Educacional)
            ml_doc = pd.DataFrame([
                {"Tópico": "O que é?", "Explicação": "Um 'Conselho de 100 Especialistas' (Random Forest) que analisa o passado."},
                {"Tópico": "O que ele olha?", "Explicação": "RSI, Volume, Tendência (ADX), Volatilidade (ATR) e Distância das Bandas."},
                {"Tópico": "Qual o objetivo?", "Explicação": "Responder: 'Qual a chance matemática desse trade dar 1% de lucro nas próximas 3h?'"},
                {"Tópico": "Notas (Score)", "Explicação": "0-40%: BLOQUEIA (Risco) | 40-80%: APROVA (Normal) | 80-100%: DIAMANTE (Alta Confiança)."},
                {"Tópico": "Status Atual", "Explicação": "O modelo treina automaticamente a cada 1 hora com novos dados da Binance."}
            ])
            ml_doc.to_excel(writer, sheet_name='📘 Entenda o ML', index=False)

            # Aba 5: 📐 Como Funciona Fibo (Educacional)
            fibo_doc = pd.DataFrame([
                {"Tópico": "Estratégia", "Explicação": "Sniper de Retração (Compra quando o preço recua para pegar impulso)."},
                {"Tópico": "Níveis de Ouro", "Explicação": "0.618 (Golden Pocket) e 0.5 (Metade do movimento). Instituições amam esses pontos."},
                {"Tópico": "Gatilho", "Explicação": "O preço deve tocar no nível 0.618 E o RSI deve estar baixo (<30-40) ao mesmo tempo."},
                {"Tópico": "Por que funciona?", "Explicação": "É uma profecia autorrealizável: tantos robôs olham isso que o preço tende a respeitar."},
                {"Tópico": "Segurança", "Explicação": "Se o preço passar direto pelo nível com força (queda livre), o bot CANCELA a compra."}
            ])
            fibo_doc.to_excel(writer, sheet_name='📐 Entenda o Fibo', index=False)

            # Aba 6: 🧠 Dados Treino ML (Auditoria)
            try:
                ml_data_path = 'data/ml_training_data.csv'
                if os.path.exists(ml_data_path):
                    df_ml = pd.read_csv(ml_data_path)
                    df_ml.to_excel(writer, sheet_name='Dados Treino ML', index=False)
            except Exception as e:
                print(f"⚠️ Erro ao adicionar aba ML: {e}")

        # --- ESTILIZAÇÃO VISUAL (PREMIUM) ---
        wb = load_workbook(OUTPUT_FILE)
        
        # Estilo Dinheiro USDT
        money_fmt = '$ #,##0.00'
        real_fmt = 'R$ #,##0.00'
        pct_fmt = '0.00%'
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4") # Azul Profissional
        center_align = Alignment(horizontal='center')

        def style_sheet(ws):
            # Formata Cabeçalho
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            # Ajusta colunas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column_letter].width = max_length + 2

        # 1. Carteira Atual
        if 'Carteira Atual' in wb.sheetnames:
            ws = wb['Carteira Atual']
            
            # Insere Resumo no Topo
            ws.insert_rows(1, 4)
            ws['A1'] = "RESUMO DA CARTEIRA"
            ws['A1'].font = Font(size=16, bold=True, color="000000")
            
            ws['A2'] = f"Saldo Total (USDT):"
            ws['B2'] = total_wallet_val
            ws['B2'].number_format = money_fmt
            ws['B2'].font = Font(bold=True)
            
            ws['A3'] = f"Saldo Total (BRL):"
            ws['B3'] = total_wallet_val * USDT_BRL_RATE
            ws['B3'].number_format = real_fmt
            ws['B3'].font = Font(bold=True)

            ws['D2'] = f"Cotação Dólar:"
            ws['E2'] = USDT_BRL_RATE
            ws['E2'].number_format = real_fmt

            # Estiliza Tabela (começa na linha 5 agora)
            for row in ws.iter_rows(min_row=5):
                for cell in row:
                    # Formatação condicional de lucro
                    if "PnL" in ws.cell(row=5, column=cell.column).value or "Lucro" in ws.cell(row=5, column=cell.column).value:
                         if isinstance(cell.value, (int, float)):
                            if cell.value > 0: cell.font = Font(color="006100") # Verde
                            elif cell.value < 0: cell.font = Font(color="9C0006") # Vermelho
            
            style_sheet(ws)

        # 2. Extrato Geral
        if 'Extrato Geral' in wb.sheetnames:
            ws = wb['Extrato Geral']
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    col_name = ws.cell(row=1, column=cell.column).value
                    if col_name and "Lucro" in col_name:
                         if isinstance(cell.value, (int, float)):
                            if cell.value > 0: cell.font = Font(color="006100") # Verde
                            elif cell.value < 0: cell.font = Font(color="9C0006") # Vermelho
            style_sheet(ws)

        wb.save(OUTPUT_FILE)
        print(f"✅ Relatório Excel gerado com sucesso: {OUTPUT_FILE}")

        # --- SINCRONIZAÇÃO GOOGLE SHEETS ---
        if HAS_GSHEETS and os.path.exists(CREDENTIALS_FILE):
             try:
                 print("☁️ Sincronizando com Google Sheets...")
                 scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
                 creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, scope)
                 client = gspread.authorize(creds)
                 sh = client.open_by_key(GSHEET_ID)

                 def update_sheet(worksheet_name, df_data):
                     try:
                         try:
                             ws = sh.worksheet(worksheet_name)
                             ws.clear()
                         except gspread.WorksheetNotFound:
                             ws = sh.add_worksheet(title=worksheet_name, rows=100, cols=20)
                         
                         # Prepara dados (converte NaN para string vazia ou 0 para JSON compliance)
                         df_data = df_data.fillna('')
                         # Converte datas para string se necessário (pandas faz isso auto no values as vezes, mas bom garantir)
                         
                         data = [df_data.columns.values.tolist()] + df_data.values.tolist()
                         ws.update(data)
                         print(f"   > Aba '{worksheet_name}' atualizada.")
                     except Exception as e:
                        print(f"   ⚠️ Erro na aba {worksheet_name}: {e}")

                 # 1. Carteira Atual
                 if not df_portfolio.empty:
                     update_sheet("Carteira Atual", df_portfolio)
                 
                 # 2. Extrato Geral
                 if not df_extrato.empty:
                     # Remove Mês para display mais limpo, ou mantém? User pediu "tudo". Mantém.
                     update_sheet("Extrato Geral", df_extrato)

                 # 3. Dados Treino ML (se existir)
                 if 'df_ml' in locals():
                     update_sheet("Dados Treino ML", df_ml)
                 
                 # 4. Performance
                 if not df_extrato.empty:
                     # Recalcula pivot, pois a lógica estava apenas no to_excel
                     pivot = df_extrato.groupby('Moeda')[['Investido ($)', 'Lucro ($)']].sum()
                     pivot['ROI (%)'] = pivot.apply(lambda x: x['Lucro ($)'] / x['Investido ($)'] if x['Investido ($)'] != 0 else 0, axis=1)
                     pivot = pivot.sort_values('Lucro ($)', ascending=False)
                     pivot.reset_index(inplace=True) # Para ter a coluna Moeda no GSheet
                     update_sheet("Performance", pivot)

                 print("✅ Sincronização Google Sheets CONCLUÍDA!")
                 
             except Exception as gs_err:
                 print(f"❌ Erro ao sincronizar Google Sheets: {gs_err}")

    except Exception as e:
        print(f"❌ Erro ao gerar relatório Excel: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    generate_report()
